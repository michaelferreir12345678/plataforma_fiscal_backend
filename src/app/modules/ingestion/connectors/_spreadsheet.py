"""Parsing de planilhas (XLSX/CSV) para os conectores baseados em arquivo.

Regras da Sprint 1B:
- ``versao_entrega`` = checksum do arquivo (``file_checksum``).
- Layout inesperado **falha explicitamente** (``SpreadsheetLayoutError``), nunca adivinha —
  crítico para o CAPAG (mudança de layout do Tesouro não pode virar dado silencioso errado).
"""

from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable
from typing import Any

from openpyxl import load_workbook

from app.core.errors import AppError


class SpreadsheetLayoutError(AppError):
    """Layout de planilha diferente do esperado — HTTP 422 (parser não adivinha)."""

    def __init__(self, detail: str) -> None:
        super().__init__(
            status=422,
            title="Layout de planilha inesperado",
            detail=detail,
            type_="urn:plataforma-fiscal:error:spreadsheet-layout",
        )


def file_checksum(raw: bytes) -> str:
    """Checksum estável do arquivo (usado como ``versao_entrega``)."""
    return hashlib.sha256(raw).hexdigest()[:16]


def read_xlsx(
    raw: bytes, *, sheet: str | None = None, header_row: int = 1
) -> list[dict[str, Any]]:
    """Lê uma planilha XLSX usando uma aba e linha de cabeçalho explícitas.

    A maioria das fontes começa na primeira linha. Algumas planilhas oficiais, como a
    CAPAG, trazem título/metadados antes do cabeçalho; ``header_row`` permite ler esse
    layout sem tentar inferir silenciosamente onde os dados começam.
    """
    if header_row < 1:
        raise SpreadsheetLayoutError("A linha de cabeçalho deve ser maior ou igual a 1.")

    workbook = load_workbook(
        io.BytesIO(raw), read_only=True, data_only=True, keep_links=False
    )
    try:
        if sheet is not None and sheet not in workbook.sheetnames:
            raise SpreadsheetLayoutError(
                f"Aba obrigatória '{sheet}' ausente. Presentes: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet] if sheet else workbook.active
        rows = worksheet.iter_rows(min_row=header_row, values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def read_csv(raw: bytes, *, delimiter: str = ";") -> list[dict[str, Any]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader]


def read_table(raw: bytes, params: dict[str, Any]) -> list[dict[str, Any]]:
    """Lê a planilha conforme ``params['formato']`` (``xlsx``|``csv``); detecta XLSX por magic."""
    formato = params.get("formato")
    if formato == "csv":
        return read_csv(raw, delimiter=params.get("delimiter", ";"))
    if formato == "xlsx" or raw[:4] == b"PK\x03\x04":
        return read_xlsx(
            raw,
            sheet=params.get("sheet"),
            header_row=int(params.get("header_row") or 1),
        )
    return read_csv(raw, delimiter=params.get("delimiter", ";"))


def require_columns(headers: Iterable[str], required: Iterable[str]) -> None:
    """Falha explicitamente se faltar alguma coluna obrigatória."""
    presentes = set(headers)
    faltando = [col for col in required if col not in presentes]
    if faltando:
        raise SpreadsheetLayoutError(
            f"Colunas obrigatórias ausentes: {faltando}. Presentes: {sorted(presentes)}"
        )
