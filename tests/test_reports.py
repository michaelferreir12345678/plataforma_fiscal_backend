"""Sprint 16: relatórios reais, fila em lote e rastreabilidade dos artefatos."""

from __future__ import annotations

import hashlib
import random
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.config import settings
from app.core.db import admin_session
from app.modules.catalog.models import DimEnte
from app.modules.indicators.models import FatoRcl
from app.modules.ingestion.models import DimEntrega
from app.modules.reports.models import Relatorio, RelatorioAgendamento
from app.modules.tenancy.models import AuditLog
from app.workers import report_tasks
from tests.conftest import auth_header, login

PERIODO = "2093-B6"
VERSAO = "reports-v1"


def _ente() -> str:
    return "7" + "".join(random.choices("0123456789", k=6))


@dataclass(frozen=True)
class ReportCase:
    entes: tuple[str, str]


@pytest.fixture
def report_case() -> Iterator[ReportCase]:
    case = ReportCase(entes=(_ente(), _ente()))
    homologada_em = datetime(2094, 1, 20, tzinfo=UTC)
    with admin_session() as session:
        for index, cod_ibge in enumerate(case.entes, start=1):
            session.add(
                DimEnte(
                    cod_ibge=cod_ibge,
                    nome=f"Município relatório {index}",
                    esfera="municipal",
                    uf="CE",
                    regiao="Nordeste",
                    populacao=100_000 + index,
                    pib=Decimal("1500000000"),
                    rpps=False,
                    possui_tcm=False,
                )
            )
            session.add(
                DimEntrega(
                    cod_ibge=cod_ibge,
                    relatorio="RREO",
                    periodo=PERIODO,
                    versao_entrega=VERSAO,
                    homologada_em=homologada_em,
                    vigente=True,
                    hash_payload=f"hash-siconfi-{cod_ibge}",
                )
            )
            session.add(
                FatoRcl(
                    cod_ibge=cod_ibge,
                    periodo_ref=PERIODO,
                    rcl_12m=Decimal("100000000") + index,
                    receita_corrente=Decimal("112000000") + index,
                    deducoes=Decimal("12000000"),
                    versao_entrega=VERSAO,
                    memoria={
                        "formula": "receita_corrente - deducoes",
                        "fonte": "SICONFI/RREO",
                    },
                )
            )
    yield case
    with admin_session() as session:
        session.execute(delete(FatoRcl).where(FatoRcl.cod_ibge.in_(case.entes)))
        session.execute(delete(DimEntrega).where(DimEntrega.cod_ibge.in_(case.entes)))
        session.execute(delete(DimEnte).where(DimEnte.cod_ibge.in_(case.entes)))


@pytest.fixture
def report_storage(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(settings, "reports_storage_dir", str(tmp_path))
    return tmp_path


@pytest.fixture
def synchronous_worker(monkeypatch: pytest.MonkeyPatch) -> None:
    def _run(ids: list) -> str:
        report_tasks.processar_lote([str(item) for item in ids])
        return "test-job"

    monkeypatch.setattr(report_tasks, "enqueue_reports", _run)


def test_modelos_e_xlsx_rastreavel_com_incompletude_explicita(
    client,
    make_org,
    report_case: ReportCase,
    report_storage: Path,
    synchronous_worker: None,
) -> None:
    org = make_org(entes=list(report_case.entes))
    token = login(client, org.email, org.senha)
    headers = auth_header(token)

    modelos = client.get("/relatorios/modelos", headers=headers)
    assert modelos.status_code == 200, modelos.text
    assert {item["codigo"] for item in modelos.json()["modelos"]} == {
        "executivo",
        "limites",
        "comparativo",
        "conformidade",
        "boletim",
    }
    assert all(item["formatos"] == ["pdf", "xlsx", "pptx"] for item in modelos.json()["modelos"])

    created = client.post(
        "/relatorios",
        headers=headers,
        json={
            "modelo": "executivo",
            "formato": "xlsx",
            "escopo": "ente",
            "ente": report_case.entes[0],
            "periodo": PERIODO,
            "as_of": "2094-02-01T00:00:00Z",
        },
    )
    assert created.status_code == 202, created.text
    assert created.json()["total_entes"] == 1
    report_id = created.json()["relatorios"][0]["id"]

    detail = client.get(f"/relatorios/{report_id}", headers=headers)
    assert detail.status_code == 200, detail.text
    body = detail.json()
    assert body["status"] == "parcial"
    assert body["gerado_em"]
    assert datetime.fromisoformat(body["as_of"]).astimezone(UTC) == datetime(
        2094, 2, 1, tzinfo=UTC
    )
    assert body["conteudo_hash"] and len(body["conteudo_hash"]) == 64
    assert body["source_refs"] == [
        {
            "relatorio": "RREO",
            "anexo": "Anexo 03",
            "periodo": PERIODO,
            "versao_entrega": VERSAO,
        }
    ] or any(source["relatorio"] == "RREO" for source in body["source_refs"])

    metrics = body["memoria"]["metricas"]
    assert {item["codigo"] for item in metrics} == {
        "rcl",
        "pessoal_executivo",
        "divida_consolidada_liquida",
        "resultado_primario",
        "saude_asps",
        "educacao_mde",
    }
    rcl = next(item for item in metrics if item["codigo"] == "rcl")
    assert Decimal(rcl["valor"]) == Decimal("100000001")
    assert datetime.fromisoformat(rcl["as_of"]).astimezone(UTC) == datetime(
        2094, 1, 20, tzinfo=UTC
    )
    assert rcl["source_refs"][0]["versao_entrega"] == VERSAO
    assert rcl["memoria"]["formula"]
    assert len(body["dados_incompletos"]) == 5
    assert all(item["tipo"] == "ausente" for item in body["dados_incompletos"])

    download = client.get(f"/relatorios/{report_id}/arquivo", headers=headers)
    assert download.status_code == 200, download.text
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert hashlib.sha256(download.content).hexdigest() == body["conteudo_hash"]
    workbook = load_workbook(BytesIO(download.content), data_only=True)
    assert {"Resumo", "Indicadores", "Pendências", "Fontes", "Memória de cálculo"}.issubset(
        set(workbook.sheetnames)
    )
    assert workbook["Indicadores"].max_row == 7
    assert workbook["Pendências"].max_row == 6
    assert workbook["Indicadores"]["J2"].value is not None  # as_of por número

    with admin_session() as session:
        actions = list(
            session.scalars(
                select(AuditLog.acao).where(
                    AuditLog.org_id == org.org_id,
                    AuditLog.recurso.like(f"%relatorio:{report_id}%"),
                )
            )
        )
    assert "EXPORTAR_RELATORIO" in actions
    assert any(report_storage.iterdir())


@pytest.mark.parametrize(
    ("formato", "signature"),
    [("pdf", b"%PDF"), ("pptx", b"PK")],
)
def test_lote_gera_um_artefato_por_ente_em_pdf_e_pptx(
    client,
    make_org,
    report_case: ReportCase,
    report_storage: Path,
    synchronous_worker: None,
    formato: str,
    signature: bytes,
) -> None:
    org = make_org(entes=list(report_case.entes))
    token = login(client, org.email, org.senha)
    headers = auth_header(token)
    response = client.post(
        "/relatorios",
        headers=headers,
        json={
            "modelo": "executivo",
            "formato": formato,
            "escopo": "lote",
            "entes": list(report_case.entes),
            "periodo": PERIODO,
            "as_of": "2094-02-01T00:00:00Z",
        },
    )
    assert response.status_code == 202, response.text
    payload = response.json()
    assert payload["total_entes"] == 2
    assert len(payload["relatorios"]) == 2
    lote_id = payload["lote_id"]

    ids = [item["id"] for item in payload["relatorios"]]
    first_detail = client.get(f"/relatorios/{ids[0]}", headers=headers).json()
    assert len(first_detail["lote_itens"]) == 2
    assert {item["cod_ibge"] for item in first_detail["lote_itens"]} == set(report_case.entes)
    assert all(item["status"] == "parcial" for item in first_detail["lote_itens"])
    for report_id in ids:
        artifact = client.get(f"/relatorios/{report_id}/arquivo", headers=headers)
        assert artifact.status_code == 200
        assert artifact.content.startswith(signature)

    with admin_session() as session:
        rows = list(
            session.scalars(
                select(Relatorio).where(
                    Relatorio.org_id == org.org_id,
                    Relatorio.lote_id == lote_id,
                )
            )
        )
    assert len(rows) == 2
    assert len({row.cod_ibge for row in rows}) == 2
    assert len(list(report_storage.iterdir())) == 2


def test_agendamento_persiste_regra_sem_gerar_antes_do_horario(
    client,
    make_org,
    report_case: ReportCase,
) -> None:
    org = make_org(entes=list(report_case.entes))
    token = login(client, org.email, org.senha)
    future = datetime.now(UTC) + timedelta(days=3)
    response = client.post(
        "/relatorios/agendamentos",
        headers=auth_header(token),
        json={
            "modelo": "boletim",
            "formato": "pdf",
            "escopo": "lote",
            "entes": list(report_case.entes),
            "periodo": PERIODO,
            "periodicidade": "mensal",
            "proxima_execucao": future.isoformat(),
        },
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["entes"] == list(report_case.entes)
    assert body["ativo"] is True

    with admin_session() as session:
        schedule = session.get(RelatorioAgendamento, body["id"])
        assert schedule is not None
        assert schedule.org_id == org.org_id
        assert schedule.parametros["modelo_versao"] == "v1"
        assert not list(session.scalars(select(Relatorio).where(Relatorio.org_id == org.org_id)))
